#!/usr/bin/env python3
"""
Fetch Leverage Edu study-location pages into a fixed-schema Excel workbook.

Default behavior:
  - First run creates the workbook and snapshot.
  - Later runs fetch fresh data, compare it with the saved snapshot, and ask
    before replacing the workbook.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import html
import json
import re
import socket
import sys
import time
from collections import OrderedDict
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SEED_URL = "https://leverageedu.com/study-locations/study-in-uk/"
DEFAULT_OUTPUT = PROJECT_ROOT / "storage" / "app" / "leverageedu_study_locations_content.xlsx"
BASE_DOMAIN = "leverageedu.com"
SCHEMA_VERSION = "2026-05-28.3"
USER_AGENT = (
    "Mozilla/5.0 (compatible; OneDegree-LeverageEduStudyLocationAudit/1.0; "
    "+https://leverageedu.com)"
)
_ORIGINAL_GETADDRINFO = socket.getaddrinfo
_DNS_FALLBACK_HOSTS: dict[str, list[str]] = {}
_DNS_FALLBACK_PATCHED = False

EXCEL_CELL_LIMIT = 32767
SAFE_CELL_LIMIT = 32000
ILLEGAL_EXCEL_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
COUNTRY_URL_RE = re.compile(r"/study-locations/study-in-[a-z0-9-]+/?$", re.I)
CF_EMAIL_RE = re.compile(r"/cdn-cgi/l/email-protection#([0-9a-fA-F]+)")
EXCLUDED_TAGS = {"script", "style", "noscript", "template", "svg"}
EXCLUDED_CONTENT_SECTION_RE = re.compile(
    r"\b(?:"
    r"what['’]?s\s+trending|trending"
    r"|top\s+universities"
    r"|top\s+medical\s+universities"
    r"|scholarships?"
    r"|student\s+visa"
    r"|frequently\s+asked"
    r"|cost\s+of\s+living"
    r"|find\s+best\s+student\s+accommodation"
    r"|still\s+exploring\s+your\s+options"
    r"|global\s+careers\s+for\s+global\s+citizens"
    r"|check\s+eligibility\s+with\s+our\s+loan"
    r")\b",
    re.I,
)

DATA_SHEETS = [
    "Pages",
    "Sections",
    "Items",
    "Links",
    "Images",
    "Raw_Text",
    "Structured_Data",
    "IndianStudents",
]

LOCATION_CHUNK_RE = re.compile(r'/_next/static/chunks/pages/study-locations/[^"\'\s]+\.js')
JSON_PARSE_RE = re.compile(r"JSON\.parse\('")
REGIONAL_FLAG_RE = re.compile(r"([\U0001F1E6-\U0001F1FF]{2})")

CONTENT_WORKBOOK_SCHEMAS: "OrderedDict[str, list[str]]" = OrderedDict(
    [
        (
            "Pages",
            [
                "country",
                "page_slug",
                "nav_label",
                "flag_code",
                "uses_eu_flag",
                "flag_alt",
                "hero_image",
                "hero_key",
                "page_title",
                "hero_heading",
                "hero_text",
                "seo_title",
                "seo_description",
            ],
        ),
        (
            "Sections",
            [
                "country",
                "page_slug",
                "section_order",
                "section_heading",
                "section_body",
            ],
        ),
        (
            "Cards",
            [
                "country",
                "page_slug",
                "section_order",
                "section_heading",
                "card_order",
                "card_title",
                "card_body",
                "cta_text",
                "cta_url",
            ],
        ),
        (
            "Courses",
            [
                "country",
                "page_slug",
                "section_order",
                "section_heading",
                "course_order",
                "university_name",
                "country_flag",
                "course_name",
                "duration",
                "credential",
                "cta_text",
                "cta_url",
            ],
        ),
        (
            "Images",
            [
                "country",
                "page_slug",
                "section_order",
                "section_heading",
                "image_order",
                "image_alt",
                "image_url",
            ],
        ),
        (
            "IndianStudents",
            [
                "country",
                "page_slug",
                "subtitle",
                "heading_before",
                "heading_highlight",
                "heading_after",
                "cta_text",
                "card_order",
                "card_value",
                "card_description",
                "card_highlighted",
            ],
        ),
        (
            "UiText",
            [
                "country",
                "page_slug",
                "text_key",
                "text_value",
            ],
        ),
    ]
)

COUNTRY_FLAG_CODES = {
    "Australia": "au",
    "Belgium": "be",
    "Canada": "ca",
    "Dubai": "ae",
    "Finland": "fi",
    "France": "fr",
    "Georgia": "ge",
    "Germany": "de",
    "Ireland": "ie",
    "Italy": "it",
    "Kazakhstan": "kz",
    "Malta": "mt",
    "Netherlands": "nl",
    "New Zealand": "nz",
    "Poland": "pl",
    "Spain": "es",
    "UK": "gb",
    "USA": "us",
}

CHANGE_REPORT_SCHEMAS: "OrderedDict[str, list[str]]" = OrderedDict(
    [
        ("Change_Summary", ["metric", "value"]),
        (
            "Pending_Changes",
            [
                "change_type",
                "content_area",
                "country",
                "field",
                "old_content",
                "new_content",
            ],
        ),
    ]
)

VOLATILE_COMPARE_FIELDS = {
    "fetched_at_utc",
    "raw_html_sha256",
    "row_sha256",
}

SHEET_SCHEMAS: "OrderedDict[str, list[str]]" = OrderedDict(
    [
        (
            "Summary",
            ["key", "value", "notes"],
        ),
        (
            "Pages",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "canonical_url",
                "http_status",
                "fetched_at_utc",
                "page_title",
                "meta_title",
                "meta_description",
                "h1",
                "hero_text",
                "section_count",
                "item_count",
                "raw_text_lines",
                "link_count",
                "image_count",
                "visible_text_sha256",
                "raw_html_sha256",
                "row_sha256",
            ],
        ),
        (
            "Sections",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "section_order",
                "section_title",
                "section_text",
                "section_text_char_count",
                "section_text_truncated",
                "item_count",
                "link_count",
                "image_count",
                "text_sha256",
                "row_sha256",
            ],
        ),
        (
            "Items",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "section_order",
                "section_title",
                "item_order",
                "item_title",
                "item_text",
                "item_text_char_count",
                "item_text_truncated",
                "links",
                "row_sha256",
            ],
        ),
        (
            "Links",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "section_order",
                "section_title",
                "link_order",
                "link_text",
                "href",
                "href_domain",
                "row_sha256",
            ],
        ),
        (
            "Images",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "section_order",
                "section_title",
                "image_order",
                "alt_text",
                "src",
                "row_sha256",
            ],
        ),
        (
            "Raw_Text",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "text_order",
                "heading_path",
                "tag",
                "text",
                "row_sha256",
            ],
        ),
        (
            "Structured_Data",
            [
                "record_id",
                "country",
                "slug",
                "source_url",
                "data_source",
                "path",
                "value_type",
                "value",
                "value_char_count",
                "value_truncated",
                "row_sha256",
            ],
        ),
        (
            "Change_Log",
            [
                "run_id",
                "run_at_utc",
                "status",
                "pages_checked",
                "records_before",
                "records_after",
                "added_records",
                "removed_records",
                "modified_records",
                "changed_percent",
                "workbook_updated",
                "notes",
            ],
        ),
        (
            "Last_Change_Detail",
            [
                "run_id",
                "change_type",
                "sheet_name",
                "record_id",
                "country",
                "field_name",
                "old_value",
                "new_value",
                "old_hash",
                "new_hash",
            ],
        ),
    ]
)


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def make_run_id() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = ILLEGAL_EXCEL_RE.sub(" ", text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s+([?.!,;:])", r"\1", text)
    return text


def clean_html_text(value: Any) -> str:
    text = html.unescape(str(value or ""))
    if "<" not in text or ">" not in text:
        return clean_text(text)
    soup = BeautifulSoup(text, "lxml")
    stripped = clean_text(soup.get_text(" "))
    stripped = re.sub(r"\(\s+", "(", stripped)
    stripped = re.sub(r"\s+\)", ")", stripped)
    return stripped


def is_excluded_content_row(row: dict[str, Any]) -> bool:
    searchable = " | ".join(
        clean_text(row.get(key, ""))
        for key in (
            "section_title",
            "section_heading",
            "heading_path",
        )
    )
    return bool(EXCLUDED_CONTENT_SECTION_RE.search(searchable))


def apply_content_exclusions(dataset: dict[str, Any]) -> None:
    sheets = dataset.get("sheets", {})
    for sheet_name in ("Sections", "Items", "Links", "Images", "Raw_Text", "Structured_Data"):
        sheets[sheet_name] = [
            row
            for row in sheets.get(sheet_name, [])
            if not is_excluded_content_row(row)
        ]
    refresh_page_summaries(dataset)


def refresh_page_summaries(dataset: dict[str, Any]) -> None:
    sheets = dataset.get("sheets", {})
    page_rows = sheets.get("Pages", [])
    slugs = [clean_text(page.get("slug", "")) for page in page_rows]
    counts = {
        slug: {
            "section_count": 0,
            "item_count": 0,
            "raw_text_lines": 0,
            "link_count": 0,
            "image_count": 0,
            "visible_text": [],
        }
        for slug in slugs
    }

    for row in sheets.get("Sections", []):
        slug = clean_text(row.get("slug", ""))
        if slug in counts:
            counts[slug]["section_count"] += 1
    for row in sheets.get("Items", []):
        slug = clean_text(row.get("slug", ""))
        if slug in counts:
            counts[slug]["item_count"] += 1
    for row in sheets.get("Raw_Text", []):
        slug = clean_text(row.get("slug", ""))
        if slug in counts:
            counts[slug]["raw_text_lines"] += 1
            counts[slug]["visible_text"].append(clean_text(row.get("text", "")))
    for row in sheets.get("Links", []):
        slug = clean_text(row.get("slug", ""))
        if slug in counts:
            counts[slug]["link_count"] += 1
    for row in sheets.get("Images", []):
        slug = clean_text(row.get("slug", ""))
        if slug in counts:
            counts[slug]["image_count"] += 1

    refreshed_pages = []
    for page in page_rows:
        slug = clean_text(page.get("slug", ""))
        summary = counts.get(slug)
        if not summary:
            refreshed_pages.append(page)
            continue

        updated = dict(page)
        updated["section_count"] = summary["section_count"]
        updated["item_count"] = summary["item_count"]
        updated["raw_text_lines"] = summary["raw_text_lines"]
        updated["link_count"] = summary["link_count"]
        updated["image_count"] = summary["image_count"]
        updated["visible_text_sha256"] = sha256_text("\n".join(summary["visible_text"]))
        updated.pop("row_sha256", None)
        refreshed_pages.append(add_hash(updated))

    sheets["Pages"] = refreshed_pages


def excel_safe(value: Any) -> str:
    text = clean_text(value)
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return f"{text[:SAFE_CELL_LIMIT]} ... [TRUNCATED original_char_count={len(text)}]"


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def id_fragment(value: str, max_length: int = 36) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", clean_text(value)).strip("_").lower()
    if not cleaned:
        cleaned = "blank"
    if len(cleaned) <= max_length:
        return cleaned
    return f"{cleaned[: max_length - 9]}_{sha256_text(cleaned)[:8]}"


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def row_hash(row: dict[str, Any]) -> str:
    comparable = {
        key: clean_text(value)
        for key, value in row.items()
        if key not in VOLATILE_COMPARE_FIELDS
    }
    return sha256_text(stable_json(comparable))


def add_hash(row: dict[str, Any]) -> dict[str, Any]:
    output = dict(row)
    output["row_sha256"] = row_hash(output)
    return output


def normalize_url(url: str, base_url: str = DEFAULT_SEED_URL) -> str:
    absolute = urljoin(base_url, url)
    parsed = urlparse(absolute)
    path = parsed.path or "/"
    if not path.endswith("/"):
        path = f"{path}/"
    return f"{parsed.scheme}://{parsed.netloc}{path}"


def absolute_url(url: str, base_url: str) -> str:
    return urljoin(base_url, url)


def decode_cloudflare_email(hex_value: str) -> str:
    try:
        key = int(hex_value[:2], 16)
        chars = [
            chr(int(hex_value[index : index + 2], 16) ^ key)
            for index in range(2, len(hex_value), 2)
        ]
        return "".join(chars)
    except ValueError:
        return ""


def content_href_url(url: str, base_url: str) -> str:
    absolute = absolute_url(url, base_url)
    match = CF_EMAIL_RE.search(absolute)
    if match:
        decoded = decode_cloudflare_email(match.group(1))
        if decoded:
            return f"mailto:{decoded}"
    return absolute


def slug_from_url(url: str) -> str:
    path = urlparse(url).path.strip("/")
    return path.split("/")[-1]


def country_from_slug(slug: str) -> str:
    country = slug.removeprefix("study-in-").replace("-", " ")
    special = {"uk": "UK", "usa": "USA", "uae": "UAE"}
    return special.get(country.lower(), country.title())


def hero_key_from_slug(slug: str) -> str:
    return slug.removeprefix("study-in-")


def page_flag_code(country: str) -> str:
    return COUNTRY_FLAG_CODES.get(country, "")


def page_uses_eu_flag(country: str) -> str:
    return "yes" if country == "Europe" else "no"


def split_hero_ctas(hero_text: str) -> tuple[str, str, str]:
    text = clean_text(hero_text)
    cta_candidates = ["Start your Journey", "Talk to an Expert"]
    found = []
    for candidate in cta_candidates:
        if re.search(re.escape(candidate), text, re.I):
            found.append(candidate)
            text = re.sub(re.escape(candidate), "", text, flags=re.I)
    return clean_text(text), (found[0] if found else ""), (found[1] if len(found) > 1 else "")


def dedupe_heading(text: str) -> str:
    text = clean_text(text)
    if not text:
        return ""
    words = text.split()
    if len(words) % 2 == 0:
        half = len(words) // 2
        left = " ".join(words[:half])
        right = " ".join(words[half:])
        if left.casefold() == right.casefold():
            return left
    midpoint = len(text) // 2
    if len(text) % 2 == 0 and text[:midpoint].casefold() == text[midpoint:].casefold():
        return text[:midpoint].strip()
    return text


def text_nodes_between(start: Tag, stop: Tag | None) -> list[str]:
    chunks: list[str] = []
    for node in start.next_elements:
        if stop is not None and node is stop:
            break
        if not isinstance(node, NavigableString):
            continue
        parent = node.parent
        if parent is None:
            continue
        if parent.name in EXCLUDED_TAGS:
            continue
        if parent.find_parent(list(EXCLUDED_TAGS)):
            continue
        text = clean_text(node)
        if text:
            chunks.append(text)
    return compact_repeated(chunks)


def compact_repeated(values: list[str]) -> list[str]:
    compacted: list[str] = []
    for value in values:
        if compacted and compacted[-1].casefold() == value.casefold():
            continue
        compacted.append(value)
    return compacted


def tags_between(start: Tag, stop: Tag | None, tag_name: str) -> list[Tag]:
    tags: list[Tag] = []
    for node in start.next_elements:
        if stop is not None and node is stop:
            break
        if isinstance(node, Tag) and node.name == tag_name:
            tags.append(node)
    return tags


def dns_failure_for_url(exc: requests.RequestException, url: str) -> bool:
    host = (urlparse(url).hostname or "").rstrip(".").lower()
    if host != BASE_DOMAIN:
        return False

    text = str(exc).lower()
    return "getaddrinfo" in text or "name resolution" in text or "failed to resolve" in text


def resolve_host_with_doh(host: str, timeout: int) -> list[str]:
    providers = [
        (
            "Cloudflare",
            "https://1.1.1.1/dns-query",
            {"Accept": "application/dns-json"},
            {"name": host, "type": "A"},
        ),
        (
            "Google",
            "https://8.8.8.8/resolve",
            {},
            {"name": host, "type": "A"},
        ),
    ]

    for provider, endpoint, headers, params in providers:
        try:
            response = requests.get(endpoint, params=params, headers=headers, timeout=min(timeout, 10))
            response.raise_for_status()
            payload = response.json()
        except requests.RequestException as exc:
            print(f"  DNS fallback via {provider} failed: {exc}", file=sys.stderr)
            continue

        addresses = []
        for answer in payload.get("Answer", []):
            if answer.get("type") != 1:
                continue
            address = str(answer.get("data", "")).strip()
            if re.match(r"^\d{1,3}(?:\.\d{1,3}){3}$", address):
                addresses.append(address)

        if addresses:
            return addresses

    return []


def install_dns_fallback(host: str, addresses: list[str]) -> None:
    global _DNS_FALLBACK_PATCHED

    normalized = host.rstrip(".").lower()
    _DNS_FALLBACK_HOSTS[normalized] = addresses

    if _DNS_FALLBACK_PATCHED:
        return

    def patched_getaddrinfo(
        query_host: str,
        port: int | str | None,
        family: int = 0,
        type: int = 0,
        proto: int = 0,
        flags: int = 0,
    ) -> list[Any]:
        key = str(query_host).rstrip(".").lower()
        fallback_addresses = _DNS_FALLBACK_HOSTS.get(key)

        if fallback_addresses and family in {socket.AF_UNSPEC, socket.AF_INET}:
            results = []
            for address in fallback_addresses:
                results.extend(_ORIGINAL_GETADDRINFO(address, port, socket.AF_INET, type, proto, flags))
            return results

        return _ORIGINAL_GETADDRINFO(query_host, port, family, type, proto, flags)

    socket.getaddrinfo = patched_getaddrinfo
    _DNS_FALLBACK_PATCHED = True


def fetch_html(session: requests.Session, url: str, timeout: int) -> tuple[str, int, str]:
    last_error: requests.RequestException | None = None
    for attempt in range(1, 4):
        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
            return response.text, response.status_code, response.url
        except requests.RequestException as exc:
            last_error = exc
            host = (urlparse(url).hostname or "").rstrip(".").lower()
            if dns_failure_for_url(exc, url) and host not in _DNS_FALLBACK_HOSTS:
                addresses = resolve_host_with_doh(host, timeout)
                if addresses:
                    install_dns_fallback(host, addresses)
                    print(
                        f"  Local DNS failed for {host}; using DNS-over-HTTPS fallback: {', '.join(addresses)}",
                        file=sys.stderr,
                    )
                    continue
            if attempt >= 3:
                break
            print(f"  Fetch failed for {url}; retrying ({attempt}/2): {exc}", file=sys.stderr)
            time.sleep(min(2 * attempt, 5))

    if last_error:
        raise last_error
    raise RuntimeError(f"Could not fetch {url}")


_JS_BACKSLASH = "\\"


def js_single_unescape(raw: str) -> str:
    """Unescape JS single-quoted string content to its runtime value."""
    out: list[str] = []
    i = 0
    length = len(raw)
    while i < length:
        ch = raw[i]
        if ch == _JS_BACKSLASH and i + 1 < length:
            nxt = raw[i + 1]
            if nxt in ("'", '"', _JS_BACKSLASH, "/"):
                out.append(nxt)
                i += 2
                continue
            if nxt == "n":
                out.append("\n"); i += 2; continue
            if nxt == "r":
                out.append("\r"); i += 2; continue
            if nxt == "t":
                out.append("\t"); i += 2; continue
            if nxt == "b":
                out.append("\b"); i += 2; continue
            if nxt == "f":
                out.append("\f"); i += 2; continue
            if nxt == "x" and i + 3 < length:
                try:
                    out.append(chr(int(raw[i + 2 : i + 4], 16))); i += 4; continue
                except ValueError:
                    pass
            if nxt == "u" and i + 5 < length:
                try:
                    out.append(chr(int(raw[i + 2 : i + 6], 16))); i += 6; continue
                except ValueError:
                    pass
            out.append(nxt)
            i += 2
        else:
            out.append(ch)
            i += 1
    return "".join(out)


def find_js_single_string_end(raw: str, start: int) -> int:
    """raw[start] is the opening single quote. Return index of closing quote, or -1."""
    i = start + 1
    while i < len(raw):
        if raw[i] == _JS_BACKSLASH:
            i += 2
            continue
        if raw[i] == "'":
            return i
        i += 1
    return -1


def fetch_indian_students_data(
    session: requests.Session,
    seed_html: str,
    seed_url: str,
    timeout: int,
) -> dict[str, dict[str, Any]]:
    """Locate the Next.js study-locations bundle and extract per-slug KPI blocks."""
    match = LOCATION_CHUNK_RE.search(seed_html)
    if not match:
        return {}
    chunk_url = urljoin(seed_url, match.group())
    try:
        chunk_html, _, _ = fetch_html(session, chunk_url, timeout)
    except requests.RequestException:
        return {}

    results: dict[str, dict[str, Any]] = {}
    for m in JSON_PARSE_RE.finditer(chunk_html):
        open_q = m.end() - 1
        close_q = find_js_single_string_end(chunk_html, open_q)
        if close_q == -1:
            continue
        raw_str = chunk_html[open_q + 1 : close_q]
        unescaped = js_single_unescape(raw_str)
        try:
            data = json.loads(unescaped)
        except json.JSONDecodeError:
            continue
        if not isinstance(data, dict):
            continue
        slug = data.get("slug")
        kpis = data.get("kpis")
        if isinstance(slug, str) and isinstance(kpis, dict) and "cards" in kpis:
            results[slug] = kpis
    return results


def indian_students_rows(
    slug: str,
    country: str,
    kpis: dict[str, Any],
) -> list[dict[str, Any]]:
    heading = kpis.get("heading", {}) if isinstance(kpis.get("heading"), dict) else {}
    cards = kpis.get("cards", [])
    rows: list[dict[str, Any]] = []
    base = {
        "country": country,
        "slug": slug,
        "subtitle": clean_html_text(kpis.get("subtitle", "")),
        "heading_before": clean_html_text(heading.get("before", "")),
        "heading_highlight": clean_html_text(heading.get("highlight", "")),
        "heading_after": clean_html_text(heading.get("after", "")),
        "cta_text": clean_html_text(kpis.get("ctaText", "")),
    }
    for index, card in enumerate(cards, 1):
        row = dict(base)
        row.update(
            {
                "record_id": f"indian|{slug}|{index:03d}",
                "card_order": index,
                "card_value": clean_html_text(card.get("value", "")),
                "card_description": clean_html_text(card.get("description", "")),
                "card_highlighted": "yes" if card.get("highlighted") else "no",
            }
        )
        rows.append(add_hash(row))
    return rows


def discover_country_urls(session: requests.Session, seed_url: str, timeout: int) -> list[str]:
    html, _, final_url = fetch_html(session, seed_url, timeout)
    soup = BeautifulSoup(html, "lxml")
    found: "OrderedDict[str, None]" = OrderedDict()
    found[normalize_url(final_url, seed_url)] = None
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "")
        normalized = normalize_url(href, final_url)
        parsed = urlparse(normalized)
        if BASE_DOMAIN not in parsed.netloc:
            continue
        if COUNTRY_URL_RE.search(parsed.path):
            found[normalized] = None
    return list(found.keys())


def get_meta(soup: BeautifulSoup, name: str) -> str:
    tag = soup.find("meta", attrs={"name": name}) or soup.find("meta", attrs={"property": name})
    return clean_text(tag.get("content")) if tag and tag.get("content") else ""


def get_canonical(soup: BeautifulSoup, source_url: str) -> str:
    tag = soup.find("link", rel=lambda value: value and "canonical" in value)
    if tag and tag.get("href"):
        return normalize_url(str(tag.get("href")), source_url)
    return source_url


def nearest_item_container(heading: Tag) -> Tag:
    fallback: Tag = heading
    for parent in heading.parents:
        if not isinstance(parent, Tag):
            continue
        if parent.name in {"main", "body", "html"}:
            break
        if parent.find("h2"):
            break
        h3_count = len(parent.find_all("h3"))
        if h3_count == 1:
            fallback = parent
        elif h3_count > 1:
            break
    return fallback


def tag_position_key(tag: Tag) -> int:
    return id(tag)


def parse_raw_text(
    main: Tag,
    country: str,
    slug: str,
    source_url: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_h1 = ""
    current_h2 = ""
    current_h3 = ""
    seen_heading_tags: set[int] = set()
    last_value = ""
    order = 0

    for node in main.descendants:
        if not isinstance(node, NavigableString):
            continue
        parent = node.parent
        if parent is None or parent.name in EXCLUDED_TAGS or parent.find_parent(list(EXCLUDED_TAGS)):
            continue
        text = clean_text(node)
        if not text:
            continue

        heading = (
            parent
            if parent.name in {"h1", "h2", "h3", "h4", "h5", "h6"}
            else parent.find_parent(["h1", "h2", "h3", "h4", "h5", "h6"])
        )
        tag_name = parent.name
        if heading is not None:
            heading_id = tag_position_key(heading)
            heading_text = dedupe_heading(heading.get_text(" ", strip=True))
            if not heading_text or heading_id in seen_heading_tags:
                continue
            seen_heading_tags.add(heading_id)
            tag_name = heading.name
            text = heading_text
            if tag_name == "h1":
                current_h1 = text
                current_h2 = ""
                current_h3 = ""
            elif tag_name == "h2":
                current_h2 = text
                current_h3 = ""
            elif tag_name == "h3":
                current_h3 = text

        if text.casefold() == last_value.casefold():
            continue
        last_value = text
        order += 1
        heading_path = " > ".join(part for part in [current_h1, current_h2, current_h3] if part)
        row = {
            "record_id": f"raw|{slug}|{order:05d}",
            "country": country,
            "slug": slug,
            "source_url": source_url,
            "text_order": order,
            "heading_path": heading_path,
            "tag": tag_name,
            "text": excel_safe(text),
        }
        rows.append(add_hash(row))
    return rows


def parse_links(
    section_start: Tag,
    section_stop: Tag | None,
    country: str,
    slug: str,
    source_url: str,
    section_order: int,
    section_title: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for anchor in tags_between(section_start, section_stop, "a"):
        text = clean_text(anchor.get_text(" ", strip=True))
        href_raw = anchor.get("href", "")
        href = content_href_url(href_raw, source_url) if href_raw else ""
        key = (text, href)
        if not text and not href:
            continue
        if key in seen:
            continue
        seen.add(key)
        parsed = urlparse(href)
        link_order = len(rows) + 1
        row = {
            "record_id": f"link|{slug}|{section_order:03d}|{link_order:04d}",
            "country": country,
            "slug": slug,
            "source_url": source_url,
            "section_order": section_order,
            "section_title": section_title,
            "link_order": link_order,
            "link_text": text,
            "href": href,
            "href_domain": parsed.netloc,
        }
        rows.append(add_hash(row))
    return rows


def parse_images(
    section_start: Tag,
    section_stop: Tag | None,
    country: str,
    slug: str,
    source_url: str,
    section_order: int,
    section_title: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for image in tags_between(section_start, section_stop, "img"):
        src_raw = image.get("src") or image.get("data-src") or image.get("data-original") or ""
        src = absolute_url(src_raw, source_url) if src_raw else ""
        alt = clean_text(image.get("alt"))
        key = (alt, src)
        if not alt and not src:
            continue
        if key in seen:
            continue
        seen.add(key)
        image_order = len(rows) + 1
        row = {
            "record_id": f"image|{slug}|{section_order:03d}|{image_order:04d}",
            "country": country,
            "slug": slug,
            "source_url": source_url,
            "section_order": section_order,
            "section_title": section_title,
            "image_order": image_order,
            "alt_text": alt,
            "src": src,
        }
        rows.append(add_hash(row))
    return rows


def parse_items(
    section_start: Tag,
    section_stop: Tag | None,
    country: str,
    slug: str,
    source_url: str,
    section_order: int,
    section_title: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for heading in tags_between(section_start, section_stop, "h3"):
        title = dedupe_heading(heading.get_text(" ", strip=True))
        if not title:
            continue
        container = nearest_item_container(heading)
        item_text = clean_text(container.get_text(" ", strip=True))
        if not item_text:
            item_text = title
        links = []
        for anchor in container.find_all("a", href=True):
            label = clean_text(anchor.get_text(" ", strip=True))
            href = content_href_url(anchor.get("href", ""), source_url)
            if label or href:
                links.append(f"{label} <{href}>".strip())
        item_order = len(rows) + 1
        row = {
            "record_id": f"item|{slug}|{section_order:03d}|{item_order:04d}",
            "country": country,
            "slug": slug,
            "source_url": source_url,
            "section_order": section_order,
            "section_title": section_title,
            "item_order": item_order,
            "item_title": title,
            "item_text": excel_safe(item_text),
            "item_text_char_count": len(item_text),
            "item_text_truncated": "yes" if len(item_text) > EXCEL_CELL_LIMIT else "no",
            "links": excel_safe(" | ".join(compact_repeated(links))),
        }
        rows.append(add_hash(row))
    return rows


def flatten_json(value: Any, prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        if not value:
            rows.append((prefix, "{}"))
        for key in sorted(value.keys()):
            path = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(flatten_json(value[key], path))
    elif isinstance(value, list):
        if not value:
            rows.append((prefix, "[]"))
        for index, item in enumerate(value):
            path = f"{prefix}[{index}]"
            rows.extend(flatten_json(item, path))
    else:
        rows.append((prefix, value))
    return rows


def parse_json_value(text: str) -> Any | None:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def parse_structured_data(
    soup: BeautifulSoup,
    country: str,
    slug: str,
    source_url: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    sources: list[tuple[str, Any]] = []

    for index, script in enumerate(soup.find_all("script", attrs={"type": "application/ld+json"}), 1):
        text = script.string or script.get_text()
        parsed = parse_json_value(text.strip()) if text else None
        if parsed is not None:
            sources.append((f"json_ld[{index}]", parsed))

    next_script = soup.find("script", id="__NEXT_DATA__")
    if next_script:
        parsed = parse_json_value((next_script.string or next_script.get_text()).strip())
        page_props = (
            parsed.get("props", {}).get("pageProps", {})
            if isinstance(parsed, dict)
            else {}
        )
        for key in ["query", "studyLocationResponsef", "studyLocationResponseS"]:
            if key in page_props:
                sources.append((f"next_pageProps.{key}", page_props[key]))

    for source_name, source_value in sources:
        for path, value in flatten_json(source_value):
            value_type = type(value).__name__
            if value is None:
                value_text = ""
            elif isinstance(value, (dict, list)):
                value_text = stable_json(value)
            else:
                value_text = str(value)
            value_text = clean_text(value_text)
            path_hash = sha256_text(f"{source_name}|{path}")[:16]
            row = {
                "record_id": f"structured|{slug}|{id_fragment(source_name)}|{path_hash}",
                "country": country,
                "slug": slug,
                "source_url": source_url,
                "data_source": source_name,
                "path": path,
                "value_type": value_type,
                "value": excel_safe(value_text),
                "value_char_count": len(value_text),
                "value_truncated": "yes" if len(value_text) > EXCEL_CELL_LIMIT else "no",
            }
            rows.append(add_hash(row))
    return rows


def parse_page(
    url: str,
    html: str,
    status_code: int,
    fetched_at: str,
) -> dict[str, list[dict[str, Any]]]:
    soup = BeautifulSoup(html, "lxml")
    main = soup.find("main") or soup.body or soup
    slug = slug_from_url(url)
    h1_tag = main.find("h1")
    h1 = dedupe_heading(h1_tag.get_text(" ", strip=True)) if h1_tag else country_from_slug(slug)
    country = clean_text(h1.removeprefix("Study in ")) or country_from_slug(slug)
    canonical_url = get_canonical(soup, url)
    page_title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else ""
    meta_description = get_meta(soup, "description") or get_meta(soup, "og:description")
    meta_title = get_meta(soup, "title") or get_meta(soup, "og:title")

    h2_tags = main.find_all("h2")
    sections: list[dict[str, Any]] = []
    all_items: list[dict[str, Any]] = []
    all_links: list[dict[str, Any]] = []
    all_images: list[dict[str, Any]] = []

    for index, h2 in enumerate(h2_tags, 1):
        next_h2 = h2_tags[index] if index < len(h2_tags) else None
        section_title = dedupe_heading(h2.get_text(" ", strip=True))
        chunks = text_nodes_between(h2, next_h2)
        section_text = clean_text(" | ".join(chunks))
        items = parse_items(h2, next_h2, country, slug, url, index, section_title)
        links = parse_links(h2, next_h2, country, slug, url, index, section_title)
        images = parse_images(h2, next_h2, country, slug, url, index, section_title)
        all_items.extend(items)
        all_links.extend(links)
        all_images.extend(images)
        row = {
            "record_id": f"section|{slug}|{index:03d}",
            "country": country,
            "slug": slug,
            "source_url": url,
            "section_order": index,
            "section_title": section_title,
            "section_text": excel_safe(section_text),
            "section_text_char_count": len(section_text),
            "section_text_truncated": "yes" if len(section_text) > EXCEL_CELL_LIMIT else "no",
            "item_count": len(items),
            "link_count": len(links),
            "image_count": len(images),
            "text_sha256": sha256_text(section_text),
        }
        sections.append(add_hash(row))

    raw_text = parse_raw_text(main, country, slug, url)
    structured_data = parse_structured_data(soup, country, slug, url)
    visible_text = clean_text(" ".join(row["text"] for row in raw_text))

    hero_text = ""
    if h1_tag:
        intro_chunks = []
        for node in h1_tag.next_elements:
            if isinstance(node, Tag) and node.name == "h2":
                break
            if isinstance(node, NavigableString):
                text = clean_text(node)
                if text and text.casefold() != h1.casefold():
                    intro_chunks.append(text)
        hero_text = clean_text(" ".join(compact_repeated(intro_chunks)))

    page_row = {
        "record_id": f"page|{slug}",
        "country": country,
        "slug": slug,
        "source_url": url,
        "canonical_url": canonical_url,
        "http_status": status_code,
        "fetched_at_utc": fetched_at,
        "page_title": page_title,
        "meta_title": meta_title,
        "meta_description": meta_description,
        "h1": h1,
        "hero_text": excel_safe(hero_text),
        "section_count": len(sections),
        "item_count": len(all_items),
        "raw_text_lines": len(raw_text),
        "link_count": len(all_links),
        "image_count": len(all_images),
        "visible_text_sha256": sha256_text(visible_text),
        "raw_html_sha256": sha256_text(html),
    }

    return {
        "Pages": [add_hash(page_row)],
        "Sections": sections,
        "Items": all_items,
        "Links": all_links,
        "Images": all_images,
        "Raw_Text": raw_text,
        "Structured_Data": structured_data,
    }


def build_dataset(args: argparse.Namespace) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
    )

    print(f"Discovering country URLs from {args.seed_url}")
    urls = discover_country_urls(session, args.seed_url, args.timeout)
    if args.max_pages:
        urls = urls[: args.max_pages]
    print(f"Discovered {len(urls)} country page(s).")

    sheets: dict[str, list[dict[str, Any]]] = {sheet: [] for sheet in DATA_SHEETS}
    fetched_at = utc_now()
    indian_students_map: dict[str, dict[str, Any]] = {}
    for index, url in enumerate(urls, 1):
        print(f"[{index}/{len(urls)}] Fetching {url}")
        html, status_code, final_url = fetch_html(session, url, args.timeout)
        if not indian_students_map:
            indian_students_map = fetch_indian_students_data(session, html, url, args.timeout)
            print(f"  Loaded Indian Students KPI data for {len(indian_students_map)} slug(s).")
        parsed = parse_page(normalize_url(final_url, url), html, status_code, fetched_at)
        for sheet_name in DATA_SHEETS:
            sheets[sheet_name].extend(parsed.get(sheet_name, []))
        slug = slug_from_url(normalize_url(final_url, url))
        if slug in indian_students_map:
            country = ""
            if parsed.get("Pages"):
                country = parsed["Pages"][0].get("country", "")
            sheets["IndianStudents"].extend(
                indian_students_rows(slug, country or country_from_slug(slug), indian_students_map[slug])
            )
        if args.delay and index < len(urls):
            time.sleep(args.delay)

    dataset = {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": fetched_at,
        "seed_url": args.seed_url,
        "country_urls": urls,
        "sheets": sheets,
    }
    apply_content_exclusions(dataset)
    dataset["snapshot_sha256"] = dataset_hash(dataset)
    return dataset


def comparable_map(dataset: dict[str, Any]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for sheet_name in DATA_SHEETS:
        for row in dataset.get("sheets", {}).get(sheet_name, []):
            record_id = row.get("record_id", "")
            key = f"{sheet_name}::{record_id}"
            output[key] = {
                field: clean_text(value)
                for field, value in row.items()
                if field not in VOLATILE_COMPARE_FIELDS
            }
    return output


def dataset_hash(dataset: dict[str, Any]) -> str:
    return sha256_text(stable_json(comparable_map(dataset)))


def summarize_records(dataset: dict[str, Any]) -> int:
    return sum(len(dataset.get("sheets", {}).get(sheet_name, [])) for sheet_name in DATA_SHEETS)


def row_country(row: dict[str, Any]) -> str:
    return clean_text(row.get("country", ""))


def compact_row_value(row: dict[str, Any]) -> str:
    preferred = [
        "country",
        "slug",
        "section_title",
        "item_title",
        "link_text",
        "href",
        "alt_text",
        "path",
        "value",
        "text",
    ]
    summary = {key: row.get(key, "") for key in preferred if row.get(key, "") not in ("", None)}
    return excel_safe(stable_json(summary) if summary else stable_json(row))


def compare_datasets(
    old_dataset: dict[str, Any] | None,
    new_dataset: dict[str, Any],
    run_id: str,
) -> dict[str, Any]:
    if not old_dataset:
        records_after = summarize_records(new_dataset)
        return {
            "has_changes": True,
            "summary": {
                "records_before": 0,
                "records_after": records_after,
                "added_records": records_after,
                "removed_records": 0,
                "modified_records": 0,
                "changed_records": records_after,
                "changed_percent": 100.0 if records_after else 0.0,
            },
            "detail_rows": [],
        }

    old_map = comparable_map(old_dataset)
    new_map = comparable_map(new_dataset)
    old_keys = set(old_map)
    new_keys = set(new_map)
    added_keys = sorted(new_keys - old_keys)
    removed_keys = sorted(old_keys - new_keys)
    modified_keys = sorted(key for key in old_keys & new_keys if old_map[key] != new_map[key])

    detail_rows: list[dict[str, Any]] = []
    for key in added_keys:
        sheet_name, record_id = key.split("::", 1)
        row = new_map[key]
        detail_rows.append(
            {
                "run_id": run_id,
                "change_type": "added",
                "sheet_name": sheet_name,
                "record_id": record_id,
                "country": row_country(row),
                "field_name": "__row__",
                "old_value": "",
                "new_value": compact_row_value(row),
                "old_hash": "",
                "new_hash": row_hash(row),
            }
        )

    for key in removed_keys:
        sheet_name, record_id = key.split("::", 1)
        row = old_map[key]
        detail_rows.append(
            {
                "run_id": run_id,
                "change_type": "removed",
                "sheet_name": sheet_name,
                "record_id": record_id,
                "country": row_country(row),
                "field_name": "__row__",
                "old_value": compact_row_value(row),
                "new_value": "",
                "old_hash": row_hash(row),
                "new_hash": "",
            }
        )

    for key in modified_keys:
        sheet_name, record_id = key.split("::", 1)
        old_row = old_map[key]
        new_row = new_map[key]
        changed_fields = sorted(
            field
            for field in set(old_row) | set(new_row)
            if clean_text(old_row.get(field, "")) != clean_text(new_row.get(field, ""))
        )
        for field in changed_fields:
            detail_rows.append(
                {
                    "run_id": run_id,
                    "change_type": "modified",
                    "sheet_name": sheet_name,
                    "record_id": record_id,
                    "country": row_country(new_row) or row_country(old_row),
                    "field_name": field,
                    "old_value": excel_safe(old_row.get(field, "")),
                    "new_value": excel_safe(new_row.get(field, "")),
                    "old_hash": sha256_text(clean_text(old_row.get(field, ""))),
                    "new_hash": sha256_text(clean_text(new_row.get(field, ""))),
                }
            )

    changed_records = len(added_keys) + len(removed_keys) + len(modified_keys)
    denominator = max(len(old_map), len(new_map), 1)
    changed_percent = round((changed_records / denominator) * 100, 2)
    summary = {
        "records_before": len(old_map),
        "records_after": len(new_map),
        "added_records": len(added_keys),
        "removed_records": len(removed_keys),
        "modified_records": len(modified_keys),
        "changed_records": changed_records,
        "changed_percent": changed_percent,
    }
    return {
        "has_changes": changed_records > 0,
        "summary": summary,
        "detail_rows": detail_rows,
    }


def load_snapshot(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def save_snapshot(
    path: Path,
    dataset: dict[str, Any],
    change_log: list[dict[str, Any]],
    last_change_detail: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    snapshot = dict(dataset)
    snapshot["change_log"] = change_log
    snapshot["last_change_detail"] = last_change_detail
    snapshot["snapshot_sha256"] = dataset_hash(dataset)
    path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")


def normalized_sheet_rows(rows: list[dict[str, Any]], headers: list[str]) -> list[list[Any]]:
    output = []
    for row in rows:
        output.append([row.get(header, "") for header in headers])
    return output


def first_link_pair(value: str) -> tuple[str, str]:
    match = re.search(r"([^|<>]+?)\s*<([^>]+)>", clean_text(value))
    if not match:
        return "", ""
    return clean_text(match.group(1)), clean_text(match.group(2))


def parse_course_item(item: dict[str, Any]) -> dict[str, str]:
    title = clean_text(item.get("item_title", ""))
    body = clean_text(item.get("item_text", ""))
    cta_text, cta_url = first_link_pair(str(item.get("links", "")))
    flag = ""
    university = ""
    details = ""

    if title and title in body:
        before, after = body.split(title, 1)
        flag_match = REGIONAL_FLAG_RE.search(before)
        if flag_match:
            flag = flag_match.group(1)
            university = clean_text(before[: flag_match.start()])
        else:
            university = clean_text(before)
        details = clean_text(after)
    else:
        details = body

    if cta_text:
        details = clean_text(re.sub(re.escape(cta_text), "", details, flags=re.I))
    details = clean_text(re.sub(r"\bView Course Details\b", "", details, flags=re.I))

    detail_parts = [clean_text(part) for part in details.split("|") if clean_text(part)]

    return {
        "university_name": university,
        "country_flag": flag,
        "course_name": title,
        "duration": detail_parts[0] if len(detail_parts) >= 1 else "",
        "credential": detail_parts[1] if len(detail_parts) >= 2 else "",
        "cta_text": cta_text,
        "cta_url": cta_url,
    }


def first_section_like(sections: list[dict[str, Any]], slug: str, needle: str) -> dict[str, Any]:
    for section in sections:
        if section.get("slug", "") != slug:
            continue
        if is_excluded_content_row(section):
            continue
        if needle.lower() in clean_text(section.get("section_title", "")).lower():
            return section
    return {}


def page_ui_rows(page: dict[str, Any], sections: list[dict[str, Any]]) -> list[dict[str, Any]]:
    country = clean_text(page.get("country", ""))
    slug = clean_text(page.get("slug", ""))
    hero_lead, primary_cta, secondary_cta = split_hero_ctas(str(page.get("hero_text", "")))

    def section_title(needle: str) -> str:
        section = first_section_like(sections, slug, needle)
        return clean_text(section.get("section_title", ""))

    text = {
        "hero_lead": hero_lead,
        "back_label": "All destinations",
        "hero_eyebrow": "Country guide",
        "primary_cta_text": primary_cta,
        "primary_cta_url": "/contact",
        "secondary_cta_text": secondary_cta,
        "secondary_cta_url": "/contact",
        "snapshot_heading": "At a glance",
        "snapshot_universities_label": "Universities",
        "snapshot_intakes_label": "Main intakes",
        "snapshot_tuition_label": "Tuition",
        "snapshot_living_cost_label": "Living cost",
        "why_eyebrow": f"Why {country}" if country else "",
        "indian_banner_aria": "Indian students highlight banner",
        "indian_subtitle_fallback": f"Indian Students in {country}" if country else "",
        "intakes_eyebrow": "Intakes",
        "intakes_feature_heading": section_title("Intakes"),
        "courses_eyebrow": "Top Courses",
        "courses_cta_fallback": "View Course Details",
        "cities_eyebrow": "Cities",
        "city_carousel_aria": "Popular cities carousel",
        "costs_eyebrow": "Costs",
        "card_cta_noise": "Apply Now|Talk to an Expert|Learn More|View Course Details",
        "intake_status_flags": "Application Opening Soon|Plan Ahead|Primary Intake",
        "tuition_snapshot_keywords": "Undergraduate|Postgraduate",
    }

    return [
        {
            "country": country,
            "page_slug": slug,
            "text_key": key,
            "text_value": excel_safe(value),
        }
        for key, value in text.items()
        if clean_text(value) != ""
    ]


def build_content_workbook_rows(dataset: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    sheets = dataset.get("sheets", {})
    rows: dict[str, list[dict[str, Any]]] = {sheet_name: [] for sheet_name in CONTENT_WORKBOOK_SCHEMAS}

    for page in sheets.get("Pages", []):
        country = page.get("country", "")
        slug = page.get("slug", "")
        hero_key = hero_key_from_slug(slug)
        flag_code = page_flag_code(country)
        rows["Pages"].append(
            {
                "country": country,
                "page_slug": slug,
                "nav_label": country,
                "flag_code": flag_code,
                "uses_eu_flag": page_uses_eu_flag(country),
                "flag_alt": f"{country} flag" if country and flag_code else "",
                "hero_image": f"assets/heroes/{hero_key}.jpg" if hero_key else "",
                "hero_key": hero_key,
                "page_title": page.get("page_title", ""),
                "hero_heading": page.get("h1", ""),
                "hero_text": page.get("hero_text", ""),
                "seo_title": page.get("meta_title", "") or page.get("page_title", ""),
                "seo_description": page.get("meta_description", ""),
            }
        )
        rows["UiText"].extend(page_ui_rows(page, sheets.get("Sections", [])))

    for section in sheets.get("Sections", []):
        if is_excluded_content_row(section):
            continue
        rows["Sections"].append(
            {
                "country": section.get("country", ""),
                "page_slug": section.get("slug", ""),
                "section_order": section.get("section_order", ""),
                "section_heading": section.get("section_title", ""),
                "section_body": section.get("section_text", ""),
            }
        )

    for item in sheets.get("Items", []):
        if is_excluded_content_row(item):
            continue
        cta_text, cta_url = first_link_pair(str(item.get("links", "")))
        is_course = "top courses" in clean_text(item.get("section_title", "")).lower()

        if is_course:
            course = parse_course_item(item)
            rows["Courses"].append(
                {
                    "country": item.get("country", ""),
                    "page_slug": item.get("slug", ""),
                    "section_order": item.get("section_order", ""),
                    "section_heading": item.get("section_title", ""),
                    "course_order": item.get("item_order", ""),
                    "university_name": course["university_name"],
                    "country_flag": course["country_flag"],
                    "course_name": course["course_name"],
                    "duration": course["duration"],
                    "credential": course["credential"],
                    "cta_text": course["cta_text"],
                    "cta_url": course["cta_url"],
                }
            )
            continue

        rows["Cards"].append(
            {
                "country": item.get("country", ""),
                "page_slug": item.get("slug", ""),
                "section_order": item.get("section_order", ""),
                "section_heading": item.get("section_title", ""),
                "card_order": item.get("item_order", ""),
                "card_title": item.get("item_title", ""),
                "card_body": item.get("item_text", ""),
                "cta_text": cta_text,
                "cta_url": cta_url,
            }
        )

    for image in sheets.get("Images", []):
        if is_excluded_content_row(image):
            continue
        rows["Images"].append(
            {
                "country": image.get("country", ""),
                "page_slug": image.get("slug", ""),
                "section_order": image.get("section_order", ""),
                "section_heading": image.get("section_title", ""),
                "image_order": image.get("image_order", ""),
                "image_alt": image.get("alt_text", ""),
                "image_url": image.get("src", ""),
            }
        )

    for kpi in sheets.get("IndianStudents", []):
        rows["IndianStudents"].append(
            {
                "country": kpi.get("country", ""),
                "page_slug": kpi.get("slug", ""),
                "subtitle": kpi.get("subtitle", ""),
                "heading_before": kpi.get("heading_before", ""),
                "heading_highlight": kpi.get("heading_highlight", ""),
                "heading_after": kpi.get("heading_after", ""),
                "cta_text": kpi.get("cta_text", ""),
                "card_order": kpi.get("card_order", ""),
                "card_value": kpi.get("card_value", ""),
                "card_description": kpi.get("card_description", ""),
                "card_highlighted": kpi.get("card_highlighted", ""),
            }
        )

    return rows


def write_rows_workbook(
    path: Path,
    schemas: "OrderedDict[str, list[str]]",
    sheet_rows: dict[str, list[dict[str, Any]]],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, headers in schemas.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for values in normalized_sheet_rows(sheet_rows.get(sheet_name, []), headers):
            ws.append(values)
        apply_sheet_style(ws, sheet_name)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_content_json(path: Path, dataset: dict[str, Any]) -> None:
    content = {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": dataset.get("generated_at_utc", ""),
        "sheets": build_content_workbook_rows(dataset),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(content, ensure_ascii=False, indent=2), encoding="utf-8")


def make_summary_rows(
    dataset: dict[str, Any],
    comparison_summary: dict[str, Any],
    status: str,
) -> list[dict[str, Any]]:
    sheets = dataset.get("sheets", {})
    pages = sheets.get("Pages", [])
    summary_pairs = [
        ("schema_version", SCHEMA_VERSION, "Workbook tabs and headers are fixed by this script."),
        ("generated_at_utc", dataset.get("generated_at_utc", ""), "UTC timestamp of the scrape."),
        ("seed_url", dataset.get("seed_url", ""), "Starting URL used for discovery."),
        ("countries_discovered", len(dataset.get("country_urls", [])), "Study-location pages discovered from the site."),
        ("pages_fetched", len(pages), "Pages successfully fetched and parsed."),
        ("sections", len(sheets.get("Sections", [])), ""),
        ("items", len(sheets.get("Items", [])), ""),
        ("links", len(sheets.get("Links", [])), ""),
        ("images", len(sheets.get("Images", [])), ""),
        ("raw_text_lines", len(sheets.get("Raw_Text", [])), ""),
        ("structured_data_rows", len(sheets.get("Structured_Data", [])), ""),
        ("data_records_total", summarize_records(dataset), "Total comparable records across data sheets."),
        ("snapshot_sha256", dataset.get("snapshot_sha256", ""), "Hash of comparable data, excluding volatile fetch timestamps."),
        ("comparison_status", status, "Initial, updated, unchanged, or report-only."),
        ("records_before", comparison_summary.get("records_before", ""), ""),
        ("records_after", comparison_summary.get("records_after", ""), ""),
        ("added_records", comparison_summary.get("added_records", ""), ""),
        ("removed_records", comparison_summary.get("removed_records", ""), ""),
        ("modified_records", comparison_summary.get("modified_records", ""), ""),
        ("changed_percent", comparison_summary.get("changed_percent", ""), "Changed logical records divided by max(old records, new records)."),
    ]
    return [{"key": key, "value": value, "notes": notes} for key, value, notes in summary_pairs]


def apply_sheet_style(ws, sheet_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    width_caps = {
        "Summary": 48,
        "Sections": 64,
        "Items": 60,
        "Raw_Text": 70,
        "Structured_Data": 70,
        "Last_Change_Detail": 60,
    }
    default_cap = width_caps.get(sheet_name, 42)
    for col_idx, column_cells in enumerate(ws.columns, 1):
        header = clean_text(ws.cell(row=1, column=col_idx).value)
        max_len = len(header)
        sample_count = 0
        for cell in column_cells:
            if cell.row == 1:
                continue
            max_len = max(max_len, min(len(clean_text(cell.value)), default_cap))
            sample_count += 1
            if sample_count >= 250:
                break
        width = min(max(max_len + 2, 12), default_cap)
        if any(token in header.lower() for token in ["url", "href", "src", "value", "text", "description"]):
            width = min(max(width, 28), default_cap)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(
    path: Path,
    dataset: dict[str, Any],
    change_log: list[dict[str, Any]],
    last_change_detail: list[dict[str, Any]],
    comparison_summary: dict[str, Any],
    status: str,
) -> None:
    write_rows_workbook(path, CONTENT_WORKBOOK_SCHEMAS, build_content_workbook_rows(dataset))


def write_change_report(path: Path, comparison: dict[str, Any]) -> None:
    summary = comparison.get("summary", {})
    summary_rows = [
        {"metric": "changed_percent", "value": summary.get("changed_percent", "")},
        {"metric": "added_records", "value": summary.get("added_records", "")},
        {"metric": "removed_records", "value": summary.get("removed_records", "")},
        {"metric": "modified_records", "value": summary.get("modified_records", "")},
        {"metric": "records_before", "value": summary.get("records_before", "")},
        {"metric": "records_after", "value": summary.get("records_after", "")},
    ]
    pending_rows = [
        {
            "change_type": detail.get("change_type", ""),
            "content_area": detail.get("sheet_name", ""),
            "country": detail.get("country", ""),
            "field": detail.get("field_name", ""),
            "old_content": detail.get("old_value", ""),
            "new_content": detail.get("new_value", ""),
        }
        for detail in comparison.get("detail_rows", [])
    ]
    write_rows_workbook(
        path,
        CHANGE_REPORT_SCHEMAS,
        {
            "Change_Summary": summary_rows,
            "Pending_Changes": pending_rows,
        },
    )


def print_change_summary(comparison: dict[str, Any], report_path: Path | None = None) -> None:
    summary = comparison["summary"]
    print(
        "Change summary: "
        f"{summary['changed_records']} changed record(s) "
        f"({summary['changed_percent']}%). "
        f"Added={summary['added_records']}, "
        f"Removed={summary['removed_records']}, "
        f"Modified={summary['modified_records']}."
    )
    details = comparison.get("detail_rows", [])
    if details:
        print("First changes:")
        for row in details[:15]:
            old_value = clean_text(row.get("old_value", ""))[:90]
            new_value = clean_text(row.get("new_value", ""))[:90]
            print(
                f"  - {row['change_type']} | {row['sheet_name']} | "
                f"{row['record_id']} | {row['field_name']} | "
                f"{old_value!r} -> {new_value!r}"
            )
        if len(details) > 15:
            print(f"  ... {len(details) - 15} more field-level change(s).")
    if report_path:
        print(f"Full pending change report: {report_path}")


def prompt_approval() -> bool:
    answer = input("Approve replacing the existing workbook and snapshot? [y/N]: ").strip().lower()
    return answer in {"y", "yes"}


def build_log_entry(
    run_id: str,
    status: str,
    pages_checked: int,
    summary: dict[str, Any],
    workbook_updated: bool,
    notes: str,
) -> dict[str, Any]:
    return {
        "run_id": run_id,
        "run_at_utc": utc_now(),
        "status": status,
        "pages_checked": pages_checked,
        "records_before": summary.get("records_before", ""),
        "records_after": summary.get("records_after", ""),
        "added_records": summary.get("added_records", ""),
        "removed_records": summary.get("removed_records", ""),
        "modified_records": summary.get("modified_records", ""),
        "changed_percent": summary.get("changed_percent", ""),
        "workbook_updated": "yes" if workbook_updated else "no",
        "notes": notes,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape Leverage Edu study-location pages into a fixed-schema Excel workbook."
    )
    parser.add_argument("--seed-url", default=DEFAULT_SEED_URL, help="Starting country URL.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Excel workbook path.")
    parser.add_argument(
        "--snapshot",
        type=Path,
        default=None,
        help="JSON snapshot path. Defaults to output path with .snapshot.json suffix.",
    )
    parser.add_argument(
        "--content-json",
        type=Path,
        default=None,
        help="Website content JSON path. Defaults to output path with .json suffix.",
    )
    parser.add_argument(
        "--report",
        type=Path,
        default=None,
        help="Pending change report path. Defaults beside the workbook.",
    )
    parser.add_argument("--approve", action="store_true", help="Auto-approve workbook replacement.")
    parser.add_argument(
        "--no-approve",
        action="store_true",
        help="Never replace the workbook; only write the pending change report.",
    )
    parser.add_argument("--force", action="store_true", help="Rewrite workbook even if data is unchanged.")
    parser.add_argument("--max-pages", type=int, default=0, help="Limit pages for testing.")
    parser.add_argument("--delay", type=float, default=0.4, help="Delay between page requests in seconds.")
    parser.add_argument("--timeout", type=int, default=30, help="HTTP timeout in seconds.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    args.seed_url = normalize_url(args.seed_url)
    output_path = args.output.resolve()
    snapshot_path = (
        args.snapshot.resolve()
        if args.snapshot
        else output_path.with_name(f"{output_path.stem}.snapshot.json")
    )
    content_json_path = (
        args.content_json.resolve()
        if args.content_json
        else output_path.with_suffix(".json")
    )
    report_path = (
        args.report.resolve()
        if args.report
        else output_path.with_name(f"{output_path.stem}_pending_changes.xlsx")
    )

    if args.approve and args.no_approve:
        print("Use either --approve or --no-approve, not both.", file=sys.stderr)
        return 2

    run_id = make_run_id()
    old_snapshot = load_snapshot(snapshot_path)
    try:
        dataset = build_dataset(args)
    except requests.RequestException as exc:
        print(
            "Could not fetch LeverageEdu source data. "
            "Check internet/DNS access from this server and try again. "
            f"Details: {exc}",
            file=sys.stderr,
        )
        return 1
    comparison = compare_datasets(old_snapshot, dataset, run_id)
    comparison["run_id"] = run_id

    old_change_log = old_snapshot.get("change_log", []) if old_snapshot else []
    pages_checked = len(dataset.get("sheets", {}).get("Pages", []))

    if not old_snapshot:
        if output_path.exists() and not args.approve:
            print(
                "A workbook already exists but no snapshot was found. "
                "To prevent an untracked overwrite, rerun with --approve or provide --snapshot.",
                file=sys.stderr,
            )
            return 1
        summary = comparison["summary"]
        change_log = old_change_log + [
            build_log_entry(
                run_id,
                "initial",
                pages_checked,
                summary,
                True,
                "Initial scrape. Workbook and snapshot created.",
            )
        ]
        write_workbook(output_path, dataset, change_log, [], summary, "initial")
        write_content_json(content_json_path, dataset)
        save_snapshot(snapshot_path, dataset, change_log, [])
        print(f"Initial workbook created: {output_path}")
        print(f"Website content JSON created: {content_json_path}")
        print(f"Snapshot created: {snapshot_path}")
        return 0

    if not comparison["has_changes"]:
        summary = comparison["summary"]
        print("No data changes detected.")
        print_change_summary(comparison)
        write_content_json(content_json_path, dataset)
        print(f"Website content JSON refreshed: {content_json_path}")
        if args.force:
            change_log = old_change_log + [
                build_log_entry(
                    run_id,
                    "unchanged_forced_rewrite",
                    pages_checked,
                    summary,
                    True,
                    "No comparable data changed; workbook rewritten because --force was used.",
                )
            ]
            last_detail = old_snapshot.get("last_change_detail", [])
            write_workbook(output_path, dataset, change_log, last_detail, summary, "unchanged")
            write_content_json(content_json_path, dataset)
            save_snapshot(snapshot_path, dataset, change_log, last_detail)
            print(f"Workbook rewritten: {output_path}")
            print(f"Website content JSON refreshed: {content_json_path}")
        return 0

    write_change_report(report_path, comparison)
    print_change_summary(comparison, report_path)

    if args.no_approve:
        print("Workbook not replaced because --no-approve was used.")
        return 0

    approved = args.approve or prompt_approval()
    if not approved:
        print("Workbook not replaced. Existing workbook and snapshot are unchanged.")
        return 0

    summary = comparison["summary"]
    change_log = old_change_log + [
        build_log_entry(
            run_id,
            "updated",
            pages_checked,
            summary,
            True,
            "Changes approved; workbook and snapshot replaced with the latest scrape.",
        )
    ]
    last_detail = comparison.get("detail_rows", [])
    write_workbook(output_path, dataset, change_log, last_detail, summary, "updated")
    write_content_json(content_json_path, dataset)
    save_snapshot(snapshot_path, dataset, change_log, last_detail)
    print(f"Workbook updated: {output_path}")
    print(f"Website content JSON updated: {content_json_path}")
    print(f"Snapshot updated: {snapshot_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
