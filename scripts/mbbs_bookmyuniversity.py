#!/usr/bin/env python3
"""
Scrape MBBS country pages from bookmyuniversity.com into a fixed-schema workbook.

Mirrors the run-pattern of leverageedu_study_locations.py:
  - First run creates the workbook and snapshot.
  - Later runs fetch fresh data, diff it against the snapshot, and ask before
    replacing the workbook. Pending changes are written to a separate report.

Captured sections per country (extracted from each `mbbsin<slug>.aspx` page):
  - "MBBS in <Country> for Indian Students"     (intro paragraphs)
  - "Advantages of studying MBBS in <Country>"  (bullet list)
  - "Eligibility Criteria & Required Documents" (bullet list)
  - "Why <Country> Is Most Popular For MBBS"    (h4 subpoints + overview facts)
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
import time
from collections import OrderedDict
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Force UTF-8 on stdout/stderr so non-ASCII content (em-dashes, smart quotes,
# emoji from source HTML) doesn't crash print() on Windows cp1252 consoles.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, OSError):
    pass


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "storage" / "app" / "mbbs_bookmyuniversity_content.xlsx"
BASE_URL = "https://www.bookmyuniversity.com"
SCHEMA_VERSION = "2026-05-28.1"
USER_AGENT = (
    "Mozilla/5.0 (compatible; OneDegree-MBBSContentAudit/1.0; +https://onedegreeadvisory.com)"
)

EXCEL_CELL_LIMIT = 32767
SAFE_CELL_LIMIT = 32000
ILLEGAL_EXCEL_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

DEFAULT_COUNTRIES: list[dict[str, str]] = [
    {"name": "Georgia", "slug": "georgia", "flag": "ge"},
    {"name": "Russia", "slug": "russia", "flag": "ru"},
    {"name": "Kazakhstan", "slug": "kazakhstan", "flag": "kz"},
    {"name": "Kyrgyzstan", "slug": "kyrgyzstan", "flag": "kg"},
    {"name": "Uzbekistan", "slug": "uzbekistan", "flag": "uz"},
]

SECTION_KEYS = OrderedDict(
    [
        ("for_indian_students", [
            # Most specific first — these have real body content.
            r"MBBS\s+in\s+.+\s+for\s+Indian\s+Students",
            r"MBBS\s+in\s+.+:\s*Overview",
            # Broader fallback (matches "Study MBBS in X 2026-27: Top..." style).
            r"Study\s+MBBS\s+(?:in\s+)?.+\s+(?:2026|2027|Overview)",
        ]),
        ("advantages", [
            r"Advantages\s+(?:and\s+Disadvantages\s+)?of\s+(?:studying\s+)?MBBS",
            r"Advantages\s+of\s+MBBS\s+in\s+",
        ]),
        ("eligibility", [
            r"Eligibility\s+Criteria",
            r"MBBS\s+in\s+.+\s+Eligibility\s+Criteria",
        ]),
        ("why_popular", [
            # Source-page variants encountered across the 5 countries.
            r"Why\s+.+?\s+Is\s+Most\s+Popular\s+For\s+MBBS",
            r"Why\s+choose\s+.+?\s+for\s+MBBS",
            r"Why\s+(?:study\s+)?MBBS\s+in\s+",
            r"Why\s+studying\s+MBBS\s+in\s+.+?\s+is\s+so\s+popular",
            r"^Why\s+[A-Za-z]+\s*\??\s*$",
            r"^About\s+[A-Za-z]+\s*\??\s*$",
        ]),
        ("admission_process", [
            r"MBBS\s+in\s+.+?\s+Admission\s+Process",
            r"^\s*Admission\s+Process\s*$",
            r"How\s+to\s+Apply\s+for\s+MBBS",
        ]),
    ]
)

DATA_SHEETS = ["Pages", "Sections", "Bullets", "Subpoints", "Facts", "AdmissionSteps"]

ADMISSION_STEP_RE = re.compile(r"^\s*Step\s+(\d+)\s*:\s*(.+?)\s*$", re.IGNORECASE)

CONTENT_WORKBOOK_SCHEMAS: "OrderedDict[str, list[str]]" = OrderedDict(
    [
        (
            "Pages",
            [
                "country",
                "page_slug",
                "page_url",
                "page_title",
                "hero_heading",
                "hero_text",
            ],
        ),
        (
            "Sections",
            [
                "country",
                "page_slug",
                "section_key",
                "section_heading",
                "section_body",
            ],
        ),
        (
            "Bullets",
            [
                "country",
                "page_slug",
                "section_key",
                "bullet_order",
                "bullet_text",
            ],
        ),
        (
            "Subpoints",
            [
                "country",
                "page_slug",
                "subpoint_order",
                "subpoint_heading",
                "subpoint_body",
            ],
        ),
        (
            "Facts",
            [
                "country",
                "page_slug",
                "fact_order",
                "fact_label",
                "fact_value",
            ],
        ),
        (
            "AdmissionSteps",
            [
                "country",
                "page_slug",
                "step_order",
                "step_title",
                "step_body",
            ],
        ),
    ]
)

CHANGE_REPORT_SCHEMAS: "OrderedDict[str, list[str]]" = OrderedDict(
    [
        ("Change_Summary", ["metric", "value"]),
        (
            "Pending_Changes",
            [
                "change_type",
                "sheet",
                "country",
                "record_id",
                "field",
                "old_content",
                "new_content",
            ],
        ),
    ]
)

VOLATILE_COMPARE_FIELDS = {"fetched_at_utc", "row_sha256"}


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def make_run_id() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = ILLEGAL_EXCEL_RE.sub(" ", text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s+([?.!,;:])", r"\1", text)
    return text


def excel_safe(value: Any) -> str:
    text = clean_text(value)
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return f"{text[:SAFE_CELL_LIMIT]} ... [TRUNCATED original_char_count={len(text)}]"


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def row_hash(row: dict[str, Any]) -> str:
    comparable = {k: clean_text(v) for k, v in row.items() if k not in VOLATILE_COMPARE_FIELDS}
    return sha256_text(stable_json(comparable))


def add_hash(row: dict[str, Any]) -> dict[str, Any]:
    out = dict(row)
    out["row_sha256"] = row_hash(out)
    return out


def heading_level(tag: Tag) -> int:
    name = (tag.name or "").lower()
    if len(name) == 2 and name[0] == "h" and name[1].isdigit():
        return int(name[1])
    return 0


def find_heading(soup: BeautifulSoup, patterns: list[str]) -> Tag | None:
    """Return the first heading that matches a pattern, with **pattern order
    taking priority over document order**. Earlier patterns are tried against
    all headings before later patterns are considered. This lets us list the
    most specific source-page heading first and fall back to broader variants.
    """
    tags = soup.find_all(["h1", "h2", "h3", "h4"])
    for pattern in patterns:
        rgx = re.compile(pattern, re.IGNORECASE)
        for tag in tags:
            text = clean_text(tag.get_text(" ", strip=True))
            if text and rgx.search(text):
                return tag
    return None


def collect_section_content(start: Tag) -> dict[str, Any]:
    """Walk forward from `start` until we hit the next heading at any level.
    Returns body text, bullet items, and any tables in between.
    """
    body_chunks: list[str] = []
    bullets: list[str] = []
    tables: list[list[list[str]]] = []

    for node in start.next_elements:
        if node is start:
            continue
        if not isinstance(node, Tag):
            continue
        if heading_level(node) > 0:
            break
        if node.name in ("script", "style", "noscript"):
            continue
        # Skip descendants of already-processed list/table blocks
        if any(parent.name in ("ul", "ol", "table") for parent in node.parents if parent is not start):
            continue
        if node.name == "ul":
            for li in node.find_all("li", recursive=False):
                t = clean_text(li.get_text(" ", strip=True))
                if t:
                    bullets.append(t)
            continue
        if node.name == "table":
            rows: list[list[str]] = []
            for tr in node.find_all("tr"):
                cells = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
                cells = [c for c in cells if c != ""]
                if cells:
                    rows.append(cells)
            if rows:
                tables.append(rows)
            continue
        if node.name in ("p", "div", "blockquote"):
            t = clean_text(node.get_text(" ", strip=True))
            if t and (not body_chunks or t not in body_chunks[-1]):
                body_chunks.append(t)

    return {
        "body": " ".join(body_chunks).strip(),
        "bullets": bullets,
        "tables": tables,
    }


def collect_subpoints(start: Tag) -> tuple[list[dict[str, Any]], list[tuple[str, str]]]:
    """For the "Why X Is Most Popular" section, collect h4 subpoints + content.
    Also returns the country facts table rows from the first subpoint.
    """
    level = heading_level(start)
    subpoints: list[dict[str, Any]] = []
    facts: list[tuple[str, str]] = []

    sib = start.find_next()
    current: dict[str, Any] | None = None
    while sib is not None:
        if not isinstance(sib, Tag):
            sib = sib.find_next()
            continue
        sib_level = heading_level(sib)
        if 0 < sib_level <= level:
            break
        if sib_level == level + 1 or (sib.name == "h4" and sib_level == 0):
            if current is not None:
                subpoints.append(current)
            current = {
                "heading": clean_text(sib.get_text(" ", strip=True)),
                "body_parts": [],
            }
            sib = sib.find_next()
            continue
        if current is None:
            sib = sib.find_next()
            continue
        if sib.name == "table":
            for tr in sib.find_all("tr"):
                cells = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
                cells = [c for c in cells if c]
                if len(cells) >= 2:
                    facts.append((cells[0], " ".join(cells[1:])))
                elif len(cells) == 1:
                    current["body_parts"].append(cells[0])
            sib = sib.find_next_sibling()
            continue
        if sib.name in ("p", "div", "blockquote"):
            t = clean_text(sib.get_text(" ", strip=True))
            if t and t not in current["body_parts"]:
                current["body_parts"].append(t)
            sib = sib.find_next_sibling()
            continue
        if sib.name == "ul":
            items = [clean_text(li.get_text(" ", strip=True)) for li in sib.find_all("li", recursive=False)]
            items = [i for i in items if i]
            if items:
                current["body_parts"].append(" | ".join(items))
            sib = sib.find_next_sibling()
            continue
        sib = sib.find_next()

    if current is not None:
        subpoints.append(current)

    cleaned = []
    for sp in subpoints:
        body = " ".join(sp["body_parts"]).strip()
        if not sp["heading"] and not body:
            continue
        cleaned.append({"heading": sp["heading"], "body": body})
    return cleaned, facts


def _li_direct_text(li: Tag) -> str:
    """Return the visible text of an <li> excluding any nested <ul>/<ol>
    descendants. Parent items like "Upload required documents:" stay separate
    from their nested children ("Passport", "10th & 12th Marksheet", ...)
    which become their own sub-steps.
    """
    parts: list[str] = []
    for child in li.children:
        if isinstance(child, Tag):
            if child.name in ("ul", "ol"):
                continue
            parts.append(child.get_text(" ", strip=True))
        elif isinstance(child, NavigableString):
            parts.append(str(child))
    return clean_text(" ".join(parts))


def collect_admission_steps(start: Tag) -> list[dict[str, Any]]:
    """For each "Step N: <title>" h2 that follows the admission-process
    heading, return its sub-steps (the <li>s in the very next <ul>/<ol>).
    Nested lists are flattened so each <li> becomes a sub-step in source
    order, deduplicated. If no list exists, fall back to <p>aragraphs that
    sit between this step heading and the next one.
    """
    # Collect every consecutive "Step N:" h2 after the admission heading.
    step_h2s: list[tuple[Tag, int, str]] = []
    for h2 in start.find_all_next("h2"):
        text = clean_text(h2.get_text(" ", strip=True))
        m = ADMISSION_STEP_RE.match(text)
        if m:
            step_h2s.append((h2, int(m.group(1)), clean_text(m.group(2))))
        else:
            break  # first non-step h2 ends the admission section

    steps: list[dict[str, Any]] = []
    seen_orders: set[int] = set()

    for index, (h2, order, title) in enumerate(step_h2s):
        if order in seen_orders or not title:
            continue
        seen_orders.add(order)

        end_node: Tag | None = step_h2s[index + 1][0] if index + 1 < len(step_h2s) else None

        # Find the first <ul>/<ol> after this h2 — but only if it appears
        # before the next step's heading (so we don't bleed across steps).
        first_list: Tag | None = None
        para_fallback: list[str] = []
        node = h2.find_next()
        while node is not None and node is not end_node:
            if isinstance(node, Tag):
                if node.name in ("ul", "ol"):
                    first_list = node
                    break
                if node.name == "p":
                    text = clean_text(node.get_text(" ", strip=True))
                    if text and not ADMISSION_STEP_RE.match(text) and text not in para_fallback:
                        para_fallback.append(text)
            node = node.find_next()

        sub_steps: list[str] = []
        if first_list is not None:
            for li in first_list.find_all("li"):
                text = _li_direct_text(li)
                if text and text not in sub_steps and not ADMISSION_STEP_RE.match(text):
                    sub_steps.append(text)
        elif para_fallback:
            sub_steps = para_fallback

        steps.append({"order": order, "title": title, "sub_steps": sub_steps})

    steps.sort(key=lambda s: s["order"])
    return steps


def parse_country_page(country: dict[str, str], html: str, page_url: str) -> dict[str, list[dict[str, Any]]]:
    soup = BeautifulSoup(html, "lxml")
    slug = country["slug"]
    name = country["name"]

    page_title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else f"MBBS in {name}"
    h1 = soup.find("h1")
    hero_heading = clean_text(h1.get_text(" ", strip=True)) if h1 else f"Study MBBS in {name}"
    meta_desc = soup.find("meta", attrs={"name": "description"})
    hero_text = clean_text(meta_desc.get("content", "")) if meta_desc and meta_desc.get("content") else ""

    sheets: dict[str, list[dict[str, Any]]] = {sheet: [] for sheet in DATA_SHEETS}
    sheets["Pages"].append(
        add_hash(
            {
                "record_id": f"page|{slug}",
                "country": name,
                "slug": slug,
                "page_url": page_url,
                "page_title": page_title,
                "hero_heading": hero_heading,
                "hero_text": excel_safe(hero_text),
            }
        )
    )

    section_order = 0
    for key, patterns in SECTION_KEYS.items():
        # Use the raw patterns. find_heading() prefers earlier patterns over
        # later ones, so the most specific variant in SECTION_KEYS wins
        # regardless of document order on the source page.
        heading = find_heading(soup, patterns)
        if heading is None:
            continue
        section_order += 1
        if key == "why_popular":
            subpoints, facts = collect_subpoints(heading)
            body = " | ".join(f"{sp['heading']}: {sp['body']}" for sp in subpoints if sp["body"])
            sheets["Sections"].append(
                add_hash(
                    {
                        "record_id": f"section|{slug}|{key}",
                        "country": name,
                        "slug": slug,
                        "section_key": key,
                        "section_order": section_order,
                        "section_heading": clean_text(heading.get_text(" ", strip=True)),
                        "section_body": excel_safe(body),
                    }
                )
            )
            for index, sp in enumerate(subpoints, 1):
                sheets["Subpoints"].append(
                    add_hash(
                        {
                            "record_id": f"subpoint|{slug}|{index:02d}",
                            "country": name,
                            "slug": slug,
                            "subpoint_order": index,
                            "subpoint_heading": sp["heading"],
                            "subpoint_body": excel_safe(sp["body"]),
                        }
                    )
                )
            for index, (label, value) in enumerate(facts, 1):
                sheets["Facts"].append(
                    add_hash(
                        {
                            "record_id": f"fact|{slug}|{index:02d}",
                            "country": name,
                            "slug": slug,
                            "fact_order": index,
                            "fact_label": label,
                            "fact_value": excel_safe(value),
                        }
                    )
                )
            continue

        if key == "admission_process":
            steps = collect_admission_steps(heading)
            body = " | ".join(f"Step {sp['order']}: {sp['title']}" for sp in steps)
            sheets["Sections"].append(
                add_hash(
                    {
                        "record_id": f"section|{slug}|{key}",
                        "country": name,
                        "slug": slug,
                        "section_key": key,
                        "section_order": section_order,
                        "section_heading": clean_text(heading.get_text(" ", strip=True)),
                        "section_body": excel_safe(body),
                    }
                )
            )
            for sp in steps:
                # Sub-steps are persisted as a " | "-delimited string so the
                # Excel cell stays a single value (and the app splits on render).
                sub_step_text = " | ".join(sp.get("sub_steps", []))
                sheets["AdmissionSteps"].append(
                    add_hash(
                        {
                            "record_id": f"step|{slug}|{sp['order']:02d}",
                            "country": name,
                            "slug": slug,
                            "step_order": sp["order"],
                            "step_title": sp["title"],
                            "step_body": excel_safe(sub_step_text),
                        }
                    )
                )
            continue

        content = collect_section_content(heading)
        sheets["Sections"].append(
            add_hash(
                {
                    "record_id": f"section|{slug}|{key}",
                    "country": name,
                    "slug": slug,
                    "section_key": key,
                    "section_order": section_order,
                    "section_heading": clean_text(heading.get_text(" ", strip=True)),
                    "section_body": excel_safe(content["body"]),
                }
            )
        )
        for index, bullet in enumerate(content["bullets"], 1):
            sheets["Bullets"].append(
                add_hash(
                    {
                        "record_id": f"bullet|{slug}|{key}|{index:02d}",
                        "country": name,
                        "slug": slug,
                        "section_key": key,
                        "bullet_order": index,
                        "bullet_text": excel_safe(bullet),
                    }
                )
            )

    return sheets


def fetch_html(session: requests.Session, url: str, timeout: int) -> tuple[str, int, str]:
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    if not response.encoding or response.encoding.lower() == "iso-8859-1":
        response.encoding = response.apparent_encoding or "utf-8"
    return response.text, response.status_code, response.url


def build_dataset(args: argparse.Namespace, countries: list[dict[str, str]]) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "en-US,en;q=0.9"})

    sheets: dict[str, list[dict[str, Any]]] = {sheet: [] for sheet in DATA_SHEETS}
    fetched_at = utc_now()
    fetched_countries: list[str] = []
    for index, country in enumerate(countries, 1):
        url = f"{BASE_URL}/mbbsin{country['slug']}.aspx"
        print(f"[{index}/{len(countries)}] Fetching {url}")
        try:
            html, status, final_url = fetch_html(session, url, args.timeout)
        except requests.RequestException as exc:
            print(f"  - skipped: {exc}")
            continue
        if "404" in final_url or status >= 400:
            print(f"  - skipped: {country['slug']} returned {status} {final_url}")
            continue
        parsed = parse_country_page(country, html, final_url)
        for sheet in DATA_SHEETS:
            sheets[sheet].extend(parsed.get(sheet, []))
        fetched_countries.append(country["slug"])
        if args.delay and index < len(countries):
            time.sleep(args.delay)

    dataset = {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": fetched_at,
        "source": BASE_URL,
        "countries_scraped": fetched_countries,
        "sheets": sheets,
    }
    dataset["snapshot_sha256"] = dataset_hash(dataset)
    return dataset


def comparable_map(dataset: dict[str, Any]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for sheet_name in DATA_SHEETS:
        for row in dataset.get("sheets", {}).get(sheet_name, []):
            record_id = row.get("record_id", "")
            key = f"{sheet_name}::{record_id}"
            output[key] = {k: clean_text(v) for k, v in row.items() if k not in VOLATILE_COMPARE_FIELDS}
    return output


def dataset_hash(dataset: dict[str, Any]) -> str:
    return sha256_text(stable_json(comparable_map(dataset)))


def summarize_records(dataset: dict[str, Any]) -> int:
    return sum(len(dataset.get("sheets", {}).get(s, [])) for s in DATA_SHEETS)


def row_country(row: dict[str, Any]) -> str:
    return clean_text(row.get("country", ""))


def compact_row_value(row: dict[str, Any]) -> str:
    preferred = ["country", "slug", "section_key", "section_heading", "bullet_text", "subpoint_heading", "fact_label", "fact_value"]
    summary = {k: row.get(k, "") for k in preferred if row.get(k, "") not in ("", None)}
    return excel_safe(stable_json(summary) if summary else stable_json(row))


def compare_datasets(old: dict[str, Any] | None, new: dict[str, Any], run_id: str) -> dict[str, Any]:
    if not old:
        records_after = summarize_records(new)
        return {
            "has_changes": True,
            "summary": {
                "records_before": 0,
                "records_after": records_after,
                "added_records": records_after,
                "removed_records": 0,
                "modified_records": 0,
                "changed_records": records_after,
                "changed_percent": 100.0 if records_after else 0.0,
            },
            "detail_rows": [],
        }

    old_map = comparable_map(old)
    new_map = comparable_map(new)
    added = sorted(set(new_map) - set(old_map))
    removed = sorted(set(old_map) - set(new_map))
    modified = sorted(k for k in set(old_map) & set(new_map) if old_map[k] != new_map[k])

    detail_rows: list[dict[str, Any]] = []
    for k in added:
        sheet, rid = k.split("::", 1)
        row = new_map[k]
        detail_rows.append(
            {
                "run_id": run_id,
                "change_type": "added",
                "sheet": sheet,
                "record_id": rid,
                "country": row_country(row),
                "field": "__row__",
                "old_value": "",
                "new_value": compact_row_value(row),
            }
        )
    for k in removed:
        sheet, rid = k.split("::", 1)
        row = old_map[k]
        detail_rows.append(
            {
                "run_id": run_id,
                "change_type": "removed",
                "sheet": sheet,
                "record_id": rid,
                "country": row_country(row),
                "field": "__row__",
                "old_value": compact_row_value(row),
                "new_value": "",
            }
        )
    for k in modified:
        sheet, rid = k.split("::", 1)
        old_row = old_map[k]
        new_row = new_map[k]
        for field in sorted(set(old_row) | set(new_row)):
            if clean_text(old_row.get(field, "")) == clean_text(new_row.get(field, "")):
                continue
            detail_rows.append(
                {
                    "run_id": run_id,
                    "change_type": "modified",
                    "sheet": sheet,
                    "record_id": rid,
                    "country": row_country(new_row) or row_country(old_row),
                    "field": field,
                    "old_value": excel_safe(old_row.get(field, "")),
                    "new_value": excel_safe(new_row.get(field, "")),
                }
            )

    changed_records = len(added) + len(removed) + len(modified)
    denom = max(len(old_map), len(new_map), 1)
    summary = {
        "records_before": len(old_map),
        "records_after": len(new_map),
        "added_records": len(added),
        "removed_records": len(removed),
        "modified_records": len(modified),
        "changed_records": changed_records,
        "changed_percent": round((changed_records / denom) * 100, 2),
    }
    return {
        "has_changes": changed_records > 0,
        "summary": summary,
        "detail_rows": detail_rows,
    }


def load_snapshot(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def save_snapshot(path: Path, dataset: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")


def normalized_sheet_rows(rows: list[dict[str, Any]], headers: list[str]) -> list[list[Any]]:
    return [[row.get(h, "") for h in headers] for row in rows]


def apply_sheet_style(ws, sheet_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    width_caps = {
        "Sections": 70,
        "Bullets": 70,
        "Subpoints": 70,
        "Facts": 50,
    }
    cap = width_caps.get(sheet_name, 42)
    for col_idx, column_cells in enumerate(ws.columns, 1):
        header = clean_text(ws.cell(row=1, column=col_idx).value)
        max_len = len(header)
        count = 0
        for cell in column_cells:
            if cell.row == 1:
                continue
            max_len = max(max_len, min(len(clean_text(cell.value)), cap))
            count += 1
            if count >= 250:
                break
        width = min(max(max_len + 2, 14), cap)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_rows_workbook(
    path: Path,
    schemas: "OrderedDict[str, list[str]]",
    sheet_rows: dict[str, list[dict[str, Any]]],
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, headers in schemas.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for values in normalized_sheet_rows(sheet_rows.get(sheet_name, []), headers):
            ws.append(values)
        apply_sheet_style(ws, sheet_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_content_rows(dataset: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    sheets = dataset.get("sheets", {})
    rows: dict[str, list[dict[str, Any]]] = {s: [] for s in CONTENT_WORKBOOK_SCHEMAS}
    for page in sheets.get("Pages", []):
        rows["Pages"].append(
            {
                "country": page.get("country", ""),
                "page_slug": page.get("slug", ""),
                "page_url": page.get("page_url", ""),
                "page_title": page.get("page_title", ""),
                "hero_heading": page.get("hero_heading", ""),
                "hero_text": page.get("hero_text", ""),
            }
        )
    for section in sheets.get("Sections", []):
        rows["Sections"].append(
            {
                "country": section.get("country", ""),
                "page_slug": section.get("slug", ""),
                "section_key": section.get("section_key", ""),
                "section_heading": section.get("section_heading", ""),
                "section_body": section.get("section_body", ""),
            }
        )
    for bullet in sheets.get("Bullets", []):
        rows["Bullets"].append(
            {
                "country": bullet.get("country", ""),
                "page_slug": bullet.get("slug", ""),
                "section_key": bullet.get("section_key", ""),
                "bullet_order": bullet.get("bullet_order", ""),
                "bullet_text": bullet.get("bullet_text", ""),
            }
        )
    for sp in sheets.get("Subpoints", []):
        rows["Subpoints"].append(
            {
                "country": sp.get("country", ""),
                "page_slug": sp.get("slug", ""),
                "subpoint_order": sp.get("subpoint_order", ""),
                "subpoint_heading": sp.get("subpoint_heading", ""),
                "subpoint_body": sp.get("subpoint_body", ""),
            }
        )
    for fact in sheets.get("Facts", []):
        rows["Facts"].append(
            {
                "country": fact.get("country", ""),
                "page_slug": fact.get("slug", ""),
                "fact_order": fact.get("fact_order", ""),
                "fact_label": fact.get("fact_label", ""),
                "fact_value": fact.get("fact_value", ""),
            }
        )
    for step in sheets.get("AdmissionSteps", []):
        rows["AdmissionSteps"].append(
            {
                "country": step.get("country", ""),
                "page_slug": step.get("slug", ""),
                "step_order": step.get("step_order", ""),
                "step_title": step.get("step_title", ""),
                "step_body": step.get("step_body", ""),
            }
        )
    return rows


def write_content_json(path: Path, dataset: dict[str, Any]) -> None:
    payload = {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": dataset.get("generated_at_utc", ""),
        "sheets": build_content_rows(dataset),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_workbook(path: Path, dataset: dict[str, Any]) -> None:
    write_rows_workbook(path, CONTENT_WORKBOOK_SCHEMAS, build_content_rows(dataset))


def write_change_report(path: Path, comparison: dict[str, Any]) -> None:
    summary = comparison.get("summary", {})
    summary_rows = [
        {"metric": "changed_percent", "value": summary.get("changed_percent", "")},
        {"metric": "added_records", "value": summary.get("added_records", "")},
        {"metric": "removed_records", "value": summary.get("removed_records", "")},
        {"metric": "modified_records", "value": summary.get("modified_records", "")},
        {"metric": "records_before", "value": summary.get("records_before", "")},
        {"metric": "records_after", "value": summary.get("records_after", "")},
    ]
    pending_rows = [
        {
            "change_type": d.get("change_type", ""),
            "sheet": d.get("sheet", ""),
            "country": d.get("country", ""),
            "record_id": d.get("record_id", ""),
            "field": d.get("field", ""),
            "old_content": d.get("old_value", ""),
            "new_content": d.get("new_value", ""),
        }
        for d in comparison.get("detail_rows", [])
    ]
    write_rows_workbook(path, CHANGE_REPORT_SCHEMAS, {"Change_Summary": summary_rows, "Pending_Changes": pending_rows})


def print_change_summary(comparison: dict[str, Any], report_path: Path | None = None) -> None:
    s = comparison["summary"]
    print(
        f"Change summary: {s['changed_records']} changed record(s) ({s['changed_percent']}%). "
        f"Added={s['added_records']}, Removed={s['removed_records']}, Modified={s['modified_records']}."
    )
    details = comparison.get("detail_rows", [])
    if details:
        print("First changes:")
        for row in details[:15]:
            old = clean_text(row.get("old_value", ""))[:90]
            new = clean_text(row.get("new_value", ""))[:90]
            print(f"  - {row['change_type']} | {row['sheet']} | {row['record_id']} | {row['field']} | {old!r} -> {new!r}")
        if len(details) > 15:
            print(f"  ... {len(details) - 15} more change(s).")
    if report_path:
        print(f"Pending change report: {report_path}")


def prompt_approval() -> bool:
    answer = input("Approve replacing the existing workbook and snapshot? [y/N]: ").strip().lower()
    return answer in {"y", "yes"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape MBBS country pages from bookmyuniversity.com.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Excel workbook path.")
    parser.add_argument("--snapshot", type=Path, default=None, help="Snapshot JSON path.")
    parser.add_argument("--content-json", type=Path, default=None, help="Content JSON path consumed by the app.")
    parser.add_argument("--report", type=Path, default=None, help="Pending change report path.")
    parser.add_argument("--approve", action="store_true", help="Auto-approve workbook replacement.")
    parser.add_argument("--no-approve", action="store_true", help="Never replace the workbook; only emit the change report.")
    parser.add_argument("--force", action="store_true", help="Rewrite workbook even if data is unchanged.")
    parser.add_argument("--delay", type=float, default=0.5, help="Delay between requests in seconds.")
    parser.add_argument("--timeout", type=int, default=30, help="HTTP timeout in seconds.")
    parser.add_argument("--country", action="append", default=None, help="Restrict to a country slug (repeatable).")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.approve and args.no_approve:
        print("Use either --approve or --no-approve, not both.", file=sys.stderr)
        return 2

    output_path = args.output.resolve()
    snapshot_path = (args.snapshot.resolve() if args.snapshot else output_path.with_name(f"{output_path.stem}.snapshot.json"))
    content_json_path = (args.content_json.resolve() if args.content_json else output_path.with_suffix(".json"))
    report_path = (args.report.resolve() if args.report else output_path.with_name(f"{output_path.stem}_pending_changes.xlsx"))

    countries = DEFAULT_COUNTRIES
    if args.country:
        wanted = {c.lower() for c in args.country}
        countries = [c for c in DEFAULT_COUNTRIES if c["slug"].lower() in wanted]
        if not countries:
            print(f"No matching countries for --country {args.country}", file=sys.stderr)
            return 2

    run_id = make_run_id()
    old_snapshot = load_snapshot(snapshot_path)
    dataset = build_dataset(args, countries)
    comparison = compare_datasets(old_snapshot, dataset, run_id)

    if not old_snapshot:
        if output_path.exists() and not args.approve:
            print("A workbook already exists but no snapshot was found. Rerun with --approve to confirm overwrite.", file=sys.stderr)
            return 1
        write_workbook(output_path, dataset)
        write_content_json(content_json_path, dataset)
        save_snapshot(snapshot_path, dataset)
        print(f"Initial workbook created: {output_path}")
        print(f"Website content JSON created: {content_json_path}")
        print(f"Snapshot created: {snapshot_path}")
        return 0

    if not comparison["has_changes"]:
        print("No data changes detected.")
        print_change_summary(comparison)
        write_content_json(content_json_path, dataset)
        print(f"Website content JSON refreshed: {content_json_path}")
        if args.force:
            write_workbook(output_path, dataset)
            save_snapshot(snapshot_path, dataset)
            print(f"Workbook rewritten (forced): {output_path}")
        return 0

    write_change_report(report_path, comparison)
    print_change_summary(comparison, report_path)

    if args.no_approve:
        print("Workbook not replaced because --no-approve was used.")
        return 0

    approved = args.approve or prompt_approval()
    if not approved:
        print("Workbook not replaced. Existing workbook and snapshot are unchanged.")
        return 0

    write_workbook(output_path, dataset)
    write_content_json(content_json_path, dataset)
    save_snapshot(snapshot_path, dataset)
    print(f"Workbook updated: {output_path}")
    print(f"Website content JSON updated: {content_json_path}")
    print(f"Snapshot updated: {snapshot_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
